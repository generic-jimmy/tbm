import csv
import io
import json
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional

from app.auth import require_auth
from app.database import db

router = APIRouter(prefix="/api/export", tags=["export"])

FIELDS = ["ts","kind","sender_name","sender_id","chat_id","chat_title",
          "content","caption","file_id","file_name","file_size","mime_type",
          "is_forwarded","fwd_from","reply_to_id","tg_storage_msg_id",
          "tg_storage_file_id"]


async def _rows(bot_hash: str, kind: Optional[str], chat_id: Optional[int]):
    return await db.get_messages(bot_hash, kind, chat_id, limit=10000)


@router.get("/csv")
async def export_csv(bot_hash: str = Query(...),
                     kind: Optional[str] = Query(None),
                     chat_id: Optional[int] = Query(None),
                     _=Depends(require_auth)):
    rows = await _rows(bot_hash, kind, chat_id)
    buf  = io.StringIO()
    w    = csv.DictWriter(buf, fieldnames=FIELDS, extrasaction="ignore")
    w.writeheader()
    w.writerows(rows)
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=messages.csv"},
    )


@router.get("/json")
async def export_json(bot_hash: str = Query(...),
                      kind: Optional[str] = Query(None),
                      chat_id: Optional[int] = Query(None),
                      _=Depends(require_auth)):
    rows = await _rows(bot_hash, kind, chat_id)
    return Response(
        content=json.dumps(rows, default=str, ensure_ascii=False, indent=2).encode(),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=messages.json"},
    )


@router.get("/xlsx")
async def export_xlsx(bot_hash: str = Query(...),
                      kind: Optional[str] = Query(None),
                      chat_id: Optional[int] = Query(None),
                      _=Depends(require_auth)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response("openpyxl not installed", status_code=500)

    rows = await _rows(bot_hash, kind, chat_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messages"

    hdr_font = Font(bold=True, color="2AABEE")
    hdr_fill = PatternFill("solid", fgColor="0D1425")
    type_fill = {
        "text": "C8E6FA", "photo": "DDD6FE", "document": "FEF3C7",
        "video": "FFEDD5", "audio": "D1FAE5", "voice": "CFFAFE",
    }

    headers = ["Timestamp","Type","Sender","User ID","Chat ID","Chat",
               "Content","Caption","File ID","File Name","Size (bytes)","MIME",
               "Forwarded","Fwd From","Reply To","Storage Msg ID","Storage File ID"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        vals = [row.get(f, "") for f in FIELDS]
        for ci, v in enumerate(vals, 1):
            ws.cell(ri, ci, v)
        fc = type_fill.get(row.get("kind", ""), "F8FAFC")
        for ci in range(1, len(FIELDS)+1):
            ws.cell(ri, ci).fill = PatternFill("solid", fgColor=fc)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    ws.column_dimensions["G"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=messages.xlsx"},
    )
